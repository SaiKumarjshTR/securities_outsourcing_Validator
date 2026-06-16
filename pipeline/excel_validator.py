#!/usr/bin/env python
"""
excel_validator.py  (v2)
══════════════════════════════════════════════════════════════════════════════
Excel → SGML Validator — Fully Deterministic (no LLM)
Thomson Reuters Securities Outsourcing | April 2026

Usage:
  python excel_validator.py output.sgm source.xlsx
  python excel_validator.py output.sgm        # structural checks only (no L1/L4)
  python excel_validator.py output.sgm source.xlsx --fixes  # show fix list

Output:
  Console report + JSON report saved as <sgm_stem>.validator.json

Scoring:
  L1  Source Fidelity       (Excel ↔ SGML)   35 pts
  L2  Structural Compliance (DTD + keying)    35 pts
  L3  Doc-Type Rules        (type-specific)   20 pts
  L4  Data Integrity        (cell content)    10 pts
  ──────────────────────────────────────────────────
  Total                                      100 pts

Decision thresholds:
  ACCEPT               ≥ 90
  ACCEPT_WITH_WARNINGS ≥ 80
  REVIEW               ≥ 70
  REJECT               < 70, or any CRITICAL issue, or L2 < 20/35

Supported document types:
  Type 1  Floating/Tracking Error Margin Rates  (POLIDOC, landscape, multi-sheet)
  Type 2  Securities Eligible for Reduced Margin (POLIDOC, portrait)
  Type 3  Foreign Exchange Spot Risk             (POLIDOC, portrait)
  Type 4  Position Limits                        (APPENDIX, landscape)
"""

import re
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any

sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)

# ─── DOC TYPE REGISTRY ───────────────────────────────────────────────────────
# Expected structural characteristics per document type.
DOC_TYPES: Dict[str, Dict[str, Any]] = {
    'type1_floating': {
        'label':      'Type 1 — Floating/Tracking Error Margin Rates',
        'root_tag':   'POLIDOC',
        'has_block2': True,
        'landscape':  True,
    },
    'type2_lserm': {
        'label':      'Type 2 — Securities Eligible for Reduced Margin (LSERM)',
        'root_tag':   'POLIDOC',
        'has_block2': True,
        'landscape':  False,
    },
    'type3_fx': {
        'label':      'Type 3 — Foreign Exchange Spot Risk',
        'root_tag':   'POLIDOC',
        'has_block2': False,
        'landscape':  False,
    },
    'type4_positions': {
        'label':      'Type 4 — Position Limits',
        'root_tag':   'APPENDIX',
        'has_block2': False,
        'landscape':  True,
    },
}


def detect_doc_type(sgml: str) -> str:
    lower = sgml.lower()
    if 'floating and tracking error margin rates' in lower:
        return 'type1_floating'
    if 'securities eligible for reduced margin' in lower or \
       'lserm' in lower.replace(' ', ''):
        return 'type2_lserm'
    if 'foreign exchange' in lower and 'spot risk' in lower:
        return 'type3_fx'
    if '<appendix' in lower or 'listed product type' in lower or \
       'position limit' in lower:
        return 'type4_positions'
    return 'unknown'


# ─── EXCEL UTILITIES ─────────────────────────────────────────────────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _find_header_row(ws) -> int:
    """Find the 1-based header row using the same heuristic as the converter."""
    for r in range(1, min(15, ws.max_row + 1)):
        vals = [ws.cell(r, c).value
                for c in range(1, min(ws.max_column + 1, 12))]
        non_empty = [v for v in vals if v is not None]
        if len(non_empty) < 2:
            continue
        if ws.cell(r, 1).value is None:
            continue
        col_a = str(ws.cell(r, 1).value or '').strip()
        if ':' in col_a:
            continue
        if any(isinstance(v, datetime) for v in non_empty):
            continue
        if ws.cell(r, 2).value is None and len(non_empty) == 1:
            continue
        return r
    return 6  # fallback


def load_excel(xlsx_path: Optional[Path]) -> dict:
    """Load all worksheets: headers + data rows per sheet."""
    if not HAS_OPENPYXL:
        return {'error': 'openpyxl not installed — run: pip install openpyxl',
                'sheets': []}
    if not xlsx_path or not xlsx_path.exists():
        return {'error': f'File not found: {xlsx_path}', 'sheets': []}
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True,
                                    data_only=True)
    except Exception as exc:
        return {'error': f'Cannot open Excel: {exc}', 'sheets': []}

    sheets = []
    for ws in wb.worksheets:
        hdr_row = _find_header_row(ws)
        headers: List[str] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(hdr_row, c).value
            if v is not None:
                headers.append(str(v))
        data_rows: List[list] = []
        for r in range(hdr_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value
                   for c in range(1, ws.max_column + 1)]
            if any(v is not None for v in row[:4]):
                data_rows.append(row)
        sheets.append({
            'name':       ws.title,
            'header_row': hdr_row,
            'headers':    headers,
            'data_rows':  data_rows,
        })
    wb.close()
    return {'sheets': sheets, 'sheet_count': len(sheets)}


# ─── SGML UTILITIES ──────────────────────────────────────────────────────────
def _strip_tags(s: str) -> str:
    return re.sub(r'<[^>]+>', '', s).strip()


def _decode_entities(s: str) -> str:
    for entity, char in [
        ('&amp;', '&'), ('&lt;', '<'), ('&gt;', '>'),
        ('&quot;', '"'), ('&apos;', "'"),
        ('&rsquo;', '\u2019'), ('&lsquo;', '\u2018'),
        ('&rdquo;', '\u201d'), ('&ldquo;', '\u201c'),
        ('&ndash;', '\u2013'), ('&mdash;', '\u2014'),
        ('&nbsp;', ' '),
    ]:
        s = s.replace(entity, char)
    return s


def _cell_text(raw: str) -> str:
    return _decode_entities(_strip_tags(raw)).strip()


def _extract_sgmltbls(sgml: str) -> List[Dict[str, Any]]:
    """Return list of {content, has_tblhead, n_cols} for each SGMLTBL."""
    tables = []
    for m in re.finditer(r'<SGMLTBL[^>]*>(.*?)</SGMLTBL>', sgml, re.DOTALL):
        content = m.group(1)
        has_head = bool(re.search(r'<TBLHEAD\b', content))
        first_cdefs = re.search(r'<TBLCDEFS[^>]*>(.*?)</TBLCDEFS>',
                                content, re.DOTALL)
        n_cols = (len(re.findall(r'<TBLCDEF\b', first_cdefs.group(1)))
                  if first_cdefs else 0)
        tables.append({'content': content, 'has_tblhead': has_head,
                       'n_cols': n_cols})
    return tables


# ─── LINE UTILITIES ───────────────────────────────────────────────────────────

def _find_line(lines: List[str], pattern: str, flags: int = 0,
               start: int = 0) -> int:
    """Return 1-based line number of first regex match, or 0."""
    rx = re.compile(pattern, flags)
    for i in range(start, len(lines)):
        if rx.search(lines[i]):
            return i + 1
    return 0


def _context_lines(lines: List[str], line_no: int, radius: int = 2) -> str:
    """Return a numbered SGML snippet centred on line_no (1-based)."""
    if line_no <= 0 or line_no > len(lines):
        return ''
    center = line_no - 1
    start  = max(0, center - radius)
    end    = min(len(lines), center + radius + 1)
    out = []
    for i in range(start, end):
        marker = '► ' if i == center else '  '
        out.append(f'{i + 1:5d} {marker}{lines[i]}')
    return '\n'.join(out)


def _line_of_offset(sgml: str, offset: int) -> int:
    """Return 1-based line number for a byte offset within sgml."""
    return sgml[:offset].count('\n') + 1


# ─── LAYER 1: SOURCE FIDELITY ────────────────────────────────────────────────
def run_l1(sgml: str, xl_data: dict, doc_type: str,
           lines: Optional[List[str]] = None) -> dict:
    """
    L1: Source Fidelity — compare Excel source against SGML output.
    Maximum: 35 pts.

    L1-A  Sheet count == BLOCK2 count             (10 pts)
    L1-B  Excel data rows == SGML TBLROW count    (10 pts)
    L1-C  Column headers preserved exactly        (10 pts)
    L1-D  Text cell content spot-check (30 cells)  (5 pts)
    """
    issues: List[dict] = []
    score = 35
    dt = DOC_TYPES.get(doc_type, {})
    _lines = lines or sgml.splitlines()

    def add(iid: str, sev: str, msg: str, detail=None,
            line: int = 0, fix: str = ''):
        nonlocal score
        score -= {'CRITICAL': 12, 'HIGH': 8, 'MEDIUM': 3, 'LOW': 1,
                  'INFO': 0}[sev]
        entry: dict = {'id': iid, 'sev': sev, 'msg': msg}
        if line:
            entry['line'] = line
            entry['context_before'] = _context_lines(_lines, line)
        if fix:
            entry['suggested_fix'] = fix
        if detail:
            entry['detail'] = detail
        issues.append(entry)

    if xl_data.get('error'):
        issues.append({'id': 'L1-00', 'sev': 'INFO',
                       'msg': f"L1 Excel checks skipped: {xl_data['error']}"})
        return {'score': 35, 'max': 35, 'issues': issues, 'skipped': True}

    sheets = xl_data.get('sheets', [])
    sheet_count = xl_data.get('sheet_count', 0)
    tables = _extract_sgmltbls(sgml)
    # Data tables have both TBLHEAD + TBLBODY (metadata tables have TBLBODY only)
    data_tables = [t for t in tables if t['has_tblhead']]

    # ── L1-A: Sheet count vs BLOCK2 count ────────────────────────────────────
    block2_count = len(re.findall(r'<BLOCK2>', sgml))
    if dt.get('has_block2'):
        if block2_count != sheet_count:
            add('L1-A', 'HIGH',
                f'Sheet count mismatch: Excel={sheet_count} sheet(s), '
                f'SGML={block2_count} BLOCK2(s) — a worksheet may be missing')

    # ── L1-B: Data row count ──────────────────────────────────────────────────
    xl_total_rows = sum(len(s['data_rows']) for s in sheets)
    sgml_data_rows = 0
    for t in data_tables:
        body_m = re.search(r'<TBLBODY[^>]*>(.*?)</TBLBODY>',
                           t['content'], re.DOTALL)
        if body_m:
            sgml_data_rows += len(re.findall(r'<TBLROW\b', body_m.group(1)))

    if xl_total_rows > 0:
        diff = abs(sgml_data_rows - xl_total_rows)
        tolerance = max(5, int(xl_total_rows * 0.01))  # 1% or min 5 rows
        if diff > tolerance:
            add('L1-B', 'HIGH',
                f'Data row count: Excel={xl_total_rows}, SGML={sgml_data_rows} '
                f'(diff={diff}, tolerance=±{tolerance})')

    # ── L1-C: Column header match ──────────────────────────────────────────────
    if sheets and data_tables:
        xl_headers = sheets[0]['headers']
        first_dt_content = data_tables[0]['content']
        head_m = re.search(r'<TBLHEAD[^>]*>(.*?)</TBLHEAD>',
                           first_dt_content, re.DOTALL)
        if head_m:
            head_rows_raw = re.findall(r'<TBLROW[^>]*>(.*?)</TBLROW>',
                                       head_m.group(1), re.DOTALL)
            if head_rows_raw:
                sgml_head_cells = [
                    _cell_text(c)
                    for c in re.findall(r'<TBLCELL[^>]*>(.*?)</TBLCELL>',
                                        head_rows_raw[0], re.DOTALL)
                ]
                mismatches = []
                for i, (xh, sh) in enumerate(
                        zip(xl_headers, sgml_head_cells), 1):
                    if xh.strip() != sh:
                        mismatches.append(
                            f'Col {i}: Excel="{xh.strip()}" vs SGML="{sh}"')
                if mismatches:
                    add('L1-C', 'HIGH',
                        f'Column header mismatch '
                        f'({len(mismatches)}/{len(xl_headers)} differ)',
                        detail=mismatches[:5])
            else:
                add('L1-C', 'MEDIUM',
                    'No TBLROW in TBLHEAD — cannot compare column headers')
        else:
            add('L1-C', 'MEDIUM',
                'No TBLHEAD found in first data table')

    # ── L1-D: Text cell content spot-check ────────────────────────────────────
    sample_vals: List[str] = []
    for sh in sheets:
        for row in sh['data_rows'][:15]:
            for cell in row:
                if cell is None:
                    continue
                val = str(cell).strip()
                # Only meaningful text values (skip short strings and pure numbers)
                if len(val) > 3 and not re.match(r'^-?[\d.,\s]+%?$', val):
                    sample_vals.append(val)
                    if len(sample_vals) >= 30:
                        break
            if len(sample_vals) >= 30:
                break
        if len(sample_vals) >= 30:
            break

    if sample_vals:
        # Entity-encode values before comparison so "A&B" matches "A&amp;B" in SGML
        def _enc(v: str) -> str:
            return v.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        not_found = [v for v in sample_vals
                     if v not in sgml and _enc(v) not in sgml]
        miss_rate = len(not_found) / len(sample_vals)
        if miss_rate > 0.30:
            add('L1-D', 'HIGH',
                f'{len(not_found)}/{len(sample_vals)} sampled text values '
                f'not found verbatim in SGML',
                detail=not_found[:5])
        elif miss_rate > 0.10:
            add('L1-D', 'MEDIUM',
                f'{len(not_found)}/{len(sample_vals)} sampled text values '
                f'not found verbatim in SGML',
                detail=not_found[:3])

    return {'score': max(0, score), 'max': 35, 'issues': issues}


# ─── LAYER 2: STRUCTURAL COMPLIANCE ─────────────────────────────────────────
def run_l2(sgml: str, doc_type: str,
           lines: Optional[List[str]] = None) -> dict:
    """
    L2: Structural compliance (DTD + keying rules).
    Maximum: 35 pts. C1–C20 checks.
    """
    issues: List[dict] = []
    deductions = 0
    _lines = lines or sgml.splitlines()
    dt = DOC_TYPES.get(doc_type, {})
    root_tag = dt.get('root_tag', 'POLIDOC')

    now = datetime.now()
    today_str = f"{now.strftime('%B')} {now.day}, {now.year}"

    first_tag_m = re.search(r'<(POLIDOC|APPENDIX)\b', sgml)
    found_root = first_tag_m.group(1) if first_tag_m else 'NONE'

    def add(iid: str, sev: str, msg: str, detail=None,
            line: int = 0, fix: str = ''):
        nonlocal deductions
        deductions += {'CRITICAL': 10, 'HIGH': 5, 'MEDIUM': 2, 'LOW': 1}[sev]
        entry: dict = {'id': iid, 'sev': sev, 'msg': msg}
        if detail:
            entry['detail'] = detail
        if line:
            entry['line'] = line
            entry['context_before'] = _context_lines(_lines, line)
        if fix:
            entry['suggested_fix'] = fix
        issues.append(entry)

    # C1: Root tag present and properly closed
    if not first_tag_m:
        add('C1', 'CRITICAL', 'No POLIDOC or APPENDIX root tag found')
    else:
        if found_root == 'POLIDOC' and '</POLIDOC>' not in sgml:
            add('C1', 'CRITICAL', 'POLIDOC not closed — </POLIDOC> missing')
        if found_root == 'APPENDIX' and '</APPENDIX>' not in sgml:
            add('C1', 'CRITICAL', 'APPENDIX not closed — </APPENDIX> missing')

    # C2: INITID must NOT appear (business rule)
    if re.search(r'INITID\s*=', sgml, re.IGNORECASE):
        add('C2', 'HIGH',
            'INITID attribute present — must be omitted per keying rules')

    # C3: LANG="EN" required on root tag
    if 'LANG="EN"' not in sgml[:500]:
        add('C3', 'MEDIUM', 'LANG="EN" missing from root tag')

    # C4: ADDDATE must be YYYYMMDD on all types.
    #     CLIPDATE and MODDATE only apply to POLIDOC root (not APPENDIX).
    attrs_to_check = (['ADDDATE', 'CLIPDATE', 'MODDATE']
                      if root_tag == 'POLIDOC' else ['ADDDATE'])
    for attr in attrs_to_check:
        m = re.search(attr + r'="([^"]*)"', sgml[:500])
        if not m:
            add('C4', 'MEDIUM',
                f'{attr} attribute missing from root tag')
        elif not re.fullmatch(r'[0-9]{8}', m.group(1)):
            add('C4', 'HIGH',
                f'{attr}="{m.group(1)}" — not YYYYMMDD format')

    # C5: <DATE> must not equal today — must come from Excel effective date
    if root_tag == 'POLIDOC':
        date_m = re.search(r'<DATE>(.*?)</DATE>', sgml)
        if not date_m:
            add('C5', 'MEDIUM', 'No <DATE> tag found in POLIDENT')
        elif date_m.group(1).strip() == today_str:
            add('C5', 'HIGH',
                f'<DATE> equals today ({today_str}) — must be the '
                f'Excel effective date, not the conversion date')

    # C6: <N> present and non-empty
    n_m = re.search(r'<N>(.*?)</N>', sgml)
    if root_tag == 'POLIDOC':
        if not n_m:
            add('C6', 'HIGH', 'No <N> tag in POLIDENT')
        elif not n_m.group(1).strip():
            add('C6', 'HIGH', '<N> tag is empty — check Excel metadata cell')

    # C7: POLIDENT <N> must match first BLOCK2 <TI>
    if root_tag == 'POLIDOC' and n_m and n_m.group(1).strip():
        b2_ti = re.search(r'<BLOCK2>\s*<TI>(.*?)</TI>', sgml, re.DOTALL)
        if b2_ti:
            n_text  = n_m.group(1).strip()
            ti_text = _strip_tags(b2_ti.group(1)).strip()
            if n_text != ti_text:
                add('C7', 'MEDIUM',
                    f'POLIDENT <N> vs first BLOCK2 <TI> mismatch:\n'
                    f'         N  = "{n_text[:70]}"\n'
                    f'         TI = "{ti_text[:70]}"')

    # C8: APPENDIX must NOT contain BLOCK2 (DTD violation)
    if doc_type == 'type4_positions' and '<BLOCK2>' in sgml:
        add('C8', 'CRITICAL', '<BLOCK2> inside <APPENDIX> — DTD violation')

    # C9: POLIDOC must have FREEFORM; BLOCK2 required for types 1 & 2
    if root_tag == 'POLIDOC':
        if '<FREEFORM>' not in sgml:
            add('C9', 'HIGH', '<FREEFORM> tag missing')
        if dt.get('has_block2') and '<BLOCK2>' not in sgml:
            add('C9', 'HIGH', '<BLOCK2> missing — body structure broken')

    # C10: Table data must exist
    cell_count = len(re.findall(r'<TBLCELL\b', sgml))
    if cell_count == 0:
        add('C10', 'CRITICAL', 'No <TBLCELL> tags — all data is missing')
    if '<TBLBODY' not in sgml:
        add('C10', 'CRITICAL', 'No <TBLBODY> tag — table structure missing')

    # C11: No "None" literals in cells (Python None leaked into output)
    none_cells = [
        (i + 1, ln.strip())
        for i, ln in enumerate(lines)
        if '<TBLCELL' in ln and '>None<' in ln
    ]
    if none_cells:
        add('C11', 'HIGH',
            f'"None" literal in {len(none_cells)} cell(s) — '
            f'first on line {none_cells[0][0]}: {none_cells[0][1][:60]}')

    # C12: No bare & in body text (must be &amp;)
    body_text = re.sub(r'<[^>]+>', '', sgml)
    bare_amps = re.findall(r'&(?![a-zA-Z#][^;\s]{0,15};)', body_text)
    if bare_amps:
        add('C12', 'HIGH',
            f'{len(bare_amps)} bare "&" character(s) not encoded as &amp;')

    # C13: TBLCDEF count must be consistent between TBLHEAD and TBLBODY
    for tbl_i, tbl_full in enumerate(
            re.findall(r'<SGMLTBL[^>]*>(.*?)</SGMLTBL>', sgml, re.DOTALL), 1):
        head_m = re.search(r'<TBLHEAD[^>]*>(.*?)</TBLHEAD>',
                           tbl_full, re.DOTALL)
        body_m = re.search(r'<TBLBODY[^>]*>(.*?)</TBLBODY>',
                           tbl_full, re.DOTALL)
        if head_m and body_m:
            h_cdefs = len(re.findall(
                r'<TBLCDEF\b', head_m.group(1).split('<TBLROWS')[0]))
            b_cdefs = len(re.findall(
                r'<TBLCDEF\b', body_m.group(1).split('<TBLROWS')[0]))
            if h_cdefs != b_cdefs:
                add('C13', 'HIGH',
                    f'Table {tbl_i}: TBLCDEF count mismatch — '
                    f'TBLHEAD={h_cdefs} vs TBLBODY={b_cdefs} '
                    f'(columns will misalign in output)')

    # C14: All TBLROW in each TBLBODY must have the same TBLCELL count
    for tbl_i, tbl_body in enumerate(
            re.findall(r'<TBLBODY[^>]*>(.*?)</TBLBODY>',
                       sgml, re.DOTALL), 1):
        rows = re.findall(r'<TBLROW[^>]*>(.*?)</TBLROW>',
                          tbl_body, re.DOTALL)
        counts = [len(re.findall(r'<TBLCELL\b', r)) for r in rows]
        if counts:
            expected = max(set(counts), key=counts.count)
            bad = [i + 1 for i, c in enumerate(counts) if c != expected]
            if bad:
                add('C14', 'HIGH',
                    f'Table {tbl_i}: {len(bad)} row(s) have wrong cell count '
                    f'(expected {expected} per row) — rows: {bad[:5]}')

    # C15: COLSTART must be sequential 1, 2, 3, …
    colstart_errors = []
    for row_i, row in enumerate(
            re.findall(r'<TBLROW[^>]*>(.*?)</TBLROW>',
                       sgml, re.DOTALL)[:100], 1):
        starts = [int(x) for x in re.findall(r'COLSTART="([0-9]+)"', row)]
        if starts and starts != list(range(1, len(starts) + 1)):
            colstart_errors.append(row_i)
    if colstart_errors:
        add('C15', 'MEDIUM',
            f'COLSTART sequence broken in {len(colstart_errors)} row(s) '
            f'— must be 1,2,3,…: first offending rows {colstart_errors[:5]}')

    # C16: No raw accented characters (must be SGML entities)
    raw_chars = sorted(set(re.findall(
        r'[éèêàâôîçùûÉÈÊÀÂÔÎÇœŒ]', body_text)))
    if raw_chars:
        add('C16', 'MEDIUM',
            f'Raw accented chars found (must use SGML entities): {raw_chars}')

    # C17: Landscape tables — RSRV markers must be present and correctly placed
    if 'ORIENT="LANDSCAP"' in sgml:
        if '<?RSRVON>' not in sgml:
            add('C17', 'HIGH',
                'Landscape table present but <?RSRVON> is missing',
                fix='Add <?RSRVON> immediately before each landscape <TABLE>')
        if '<?RSRVOFF>' not in sgml:
            add('C17', 'HIGH',
                'Landscape table present but <?RSRVOFF> is missing',
                fix='Add <?RSRVOFF> immediately after each landscape </TABLE>')
        if not re.search(r'<\?RSRVON>\s*<TABLE>', sgml):
            ln = _find_line(_lines, r'<\?RSRVON>')
            add('C17', 'MEDIUM',
                '<?RSRVON> not immediately before <TABLE> — wrong position',
                line=ln,
                fix='Move <?RSRVON> to the line immediately before <TABLE>')

    # ── C18: Tag balance — structural tags must be properly closed ─────────────
    STRUCTURAL_TAGS = [
        'POLIDOC', 'APPENDIX', 'BLOCK1', 'BLOCK2', 'FREEFORM',
        'POLIDENT', 'SGMLTBL', 'TABLE', 'TBLHEAD', 'TBLBODY', 'TBLROW',
    ]
    for tag in STRUCTURAL_TAGS:
        opens  = len(re.findall(rf'<{tag}\b', sgml))
        closes = len(re.findall(rf'</{tag}>', sgml))
        if opens != closes:
            ln = _find_line(_lines, rf'<{tag}\b')
            add('C18', 'CRITICAL',
                f'<{tag}> not balanced: {opens} open vs {closes} close tag(s)',
                line=ln,
                fix=f'Add {abs(opens - closes)} missing '
                    f'{'</' + tag + '>' if opens > closes else '<' + tag + '>'}'
                    f' tag(s)')

    # ── C19: TBLROWS attribute must match actual TBLROW count ──────────────────
    for idx, m in enumerate(
            re.finditer(r'<TBLBODY(\b[^>]*)>(.*?)</TBLBODY>',
                        sgml, re.DOTALL), 1):
        attrs_str    = m.group(1)
        body_content = m.group(2)
        actual_rows  = len(re.findall(r'<TBLROW\b', body_content))
        tblrows_m    = re.search(r'TBLROWS="(\d+)"', attrs_str)
        ln = _line_of_offset(sgml, m.start())
        if tblrows_m:
            declared = int(tblrows_m.group(1))
            if declared != actual_rows:
                add('C19', 'MEDIUM',
                    f'TBLBODY #{idx}: TBLROWS="{declared}" declared but '
                    f'{actual_rows} actual <TBLROW> tags',
                    line=ln,
                    fix=f'Change TBLROWS="{declared}" to TBLROWS="{actual_rows}"')
        elif actual_rows > 0:
            add('C19', 'LOW',
                f'TBLBODY #{idx}: TBLROWS attribute missing ({actual_rows} rows)',
                line=ln,
                fix=f'Add TBLROWS="{actual_rows}" to the <TBLBODY> opening tag')

    # ── C20: Landscape SGMLTBL must carry a TBLWD attribute ───────────────────
    for m in re.finditer(r'<SGMLTBL\b([^>]*)>', sgml):
        if 'LANDSCAP' in m.group(1) and 'TBLWD=' not in m.group(1):
            ln = _line_of_offset(sgml, m.start())
            add('C20', 'MEDIUM',
                f'Landscape <SGMLTBL> on line {ln} missing TBLWD attribute',
                line=ln,
                fix='Add TBLWD="100" (or appropriate width) to '
                    '<SGMLTBL ORIENT="LANDSCAP" ...>')

    score = max(0, 35 - deductions)
    return {'score': score, 'max': 35, 'issues': issues,
            'cell_count': cell_count}


# ─── LAYER 3: DOC-TYPE COMPLIANCE ────────────────────────────────────────────
def run_l3(sgml: str, doc_type: str,
           lines: Optional[List[str]] = None) -> dict:
    """
    L3: Document-type specific compliance rules.
    Maximum: 20 pts.

    L3-A  Doc type identifiable from content
    L3-B  Root tag matches expected type (POLIDOC vs APPENDIX)
    L3-C  All TBLCDEFS column widths sum to 100%
    L3-D  Landscape orientation present/absent as required
    L3-E  BLOCK2 presence matches type expectation
    """
    issues: List[dict] = []
    score = 20
    dt = DOC_TYPES.get(doc_type, {})
    _lines = lines or sgml.splitlines()

    def add(iid: str, sev: str, msg: str, detail=None,
            line: int = 0, fix: str = ''):
        nonlocal score
        score -= {'CRITICAL': 8, 'HIGH': 4, 'MEDIUM': 2, 'LOW': 1}[sev]
        entry: dict = {'id': iid, 'sev': sev, 'msg': msg}
        if detail:
            entry['detail'] = detail
        if line:
            entry['line'] = line
            entry['context_before'] = _context_lines(_lines, line)
        if fix:
            entry['suggested_fix'] = fix
        issues.append(entry)

    # L3-A: Doc type must be identifiable
    if doc_type == 'unknown':
        add('L3-A', 'HIGH',
            'File type could not be determined — '
            'type-specific checks cannot be applied')
        return {'score': max(0, score), 'max': 20, 'issues': issues,
                'doc_type': doc_type, 'doc_type_label': 'Unknown'}

    # L3-B: Root tag must match expected for this type
    expected_root = dt.get('root_tag', 'POLIDOC')
    if expected_root == 'APPENDIX' and \
            not re.search(r'<APPENDIX\b', sgml, re.IGNORECASE):
        add('L3-B', 'CRITICAL',
            f'{doc_type} requires <APPENDIX> root tag, '
            f'not <POLIDOC>')
    elif expected_root == 'POLIDOC' and \
            not re.search(r'<POLIDOC\b', sgml, re.IGNORECASE):
        add('L3-B', 'CRITICAL',
            f'{doc_type} requires <POLIDOC> root tag, '
            f'not <APPENDIX>')

    # L3-C: Column widths must sum to 100% in every TBLCDEFS block
    colwidth_errors: List[str] = []
    for idx, m in enumerate(
            re.finditer(r'<TBLCDEFS[^>]*>(.*?)</TBLCDEFS>',
                        sgml, re.DOTALL | re.IGNORECASE), 1):
        widths = [int(w)
                  for w in re.findall(r'COLWD\s*=\s*"?(\d+)"?', m.group(0))]
        if not widths:
            continue
        total = sum(widths)
        if not 95 <= total <= 105:
            ctx_before = sgml[max(0, m.start() - 200):m.start()]
            section = ('TBLHEAD' if 'TBLHEAD' in ctx_before else
                       'TBLBODY' if 'TBLBODY' in ctx_before else '?')
            colwidth_errors.append(
                f'TBLCDEFS #{idx} ({section}): sum={total}, widths={widths}')
    if colwidth_errors:
        add('L3-C', 'HIGH',
            f'Column widths do not sum to 100% '
            f'in {len(colwidth_errors)} TBLCDEFS block(s)',
            detail=colwidth_errors[:3])

    # L3-D: Landscape orientation must match type expectation
    needs_landscape = dt.get('landscape', False)
    has_landscap    = 'ORIENT="LANDSCAP"' in sgml
    if needs_landscape and not has_landscap:
        add('L3-D', 'HIGH',
            f'{doc_type} requires ORIENT="LANDSCAP" on <SGMLTBL> — not found')
    elif not needs_landscape and has_landscap:
        add('L3-D', 'MEDIUM',
            f'{doc_type} should not have ORIENT="LANDSCAP" '
            f'(portrait layout expected)')

    # L3-E: BLOCK2 presence must match type expectation
    has_block2 = '<BLOCK2>' in sgml
    if dt.get('has_block2') and not has_block2:
        add('L3-E', 'HIGH',
            f'{doc_type} must have <BLOCK2> sections — none found')
    elif not dt.get('has_block2') and has_block2:
        add('L3-E', 'MEDIUM',
            f'{doc_type} should not have <BLOCK2> sections')

    return {
        'score':          max(0, score),
        'max':            20,
        'issues':         issues,
        'doc_type':       doc_type,
        'doc_type_label': dt.get('label', doc_type),
    }


_DATE_MONTHS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]
_DATE_RE = re.compile(
    r'^(' + '|'.join(_DATE_MONTHS) + r')\s+\d{1,2},\s+\d{4}$'
)


# ─── LAYER 4: DATA INTEGRITY ─────────────────────────────────────────────────
def run_l4(sgml: str, xl_data: dict, doc_type: str,
           lines: Optional[List[str]] = None) -> dict:
    """
    L4: Cell-level data integrity checks.
    Maximum: 10 pts.  L4-A through L4-F.
    """
    issues: List[dict] = []
    score = 10
    sheets = xl_data.get('sheets', []) if not xl_data.get('error') else []
    _lines = lines or sgml.splitlines()

    def add(iid: str, sev: str, msg: str, detail=None,
            line: int = 0, fix: str = ''):
        nonlocal score
        score -= {'CRITICAL': 5, 'HIGH': 3, 'MEDIUM': 1, 'LOW': 0}[sev]
        entry: dict = {'id': iid, 'sev': sev, 'msg': msg}
        if detail:
            entry['detail'] = detail
        if line:
            entry['line'] = line
            entry['context_before'] = _context_lines(_lines, line)
        if fix:
            entry['suggested_fix'] = fix
        issues.append(entry)

    # L4-A: No "None" literal values in TBLCELL
    none_count = len(re.findall(
        r'<TBLCELL[^>]*>None</TBLCELL>', sgml, re.IGNORECASE))
    if none_count:
        add('L4-A', 'HIGH',
            f'{none_count} cell(s) contain literal "None" '
            f'— Python None leaked into SGML output')

    # L4-B: Empty TBLROW blocks must not exist
    empty_rows = len(re.findall(
        r'<TBLROW[^>]*>\s*</TBLROW>', sgml, re.IGNORECASE))
    if empty_rows:
        add('L4-B', 'MEDIUM',
            f'{empty_rows} empty <TBLROW></TBLROW> block(s) '
            f'— converter must skip empty rows')

    # L4-C: Detect raw decimal fractions where percentages expected
    # e.g. 0.05 in SGML when Excel cell was formatted as 5.00%
    if sheets:
        raw_decimals: List[str] = []
        for sh in sheets:
            for row in sh['data_rows'][:50]:
                for cell in row:
                    if isinstance(cell, float) and 0 < abs(cell) < 1:
                        cell_str = repr(cell)
                        if cell_str in sgml:
                            raw_decimals.append(cell_str)
        if len(raw_decimals) > 3:
            add('L4-C', 'MEDIUM',
                f'{len(raw_decimals)} value(s) appear as raw decimal fraction '
                f'(e.g. 0.05 instead of 5.00%) — check percentage formatting',
                detail=raw_decimals[:5])

    # L4-D: Uppercase Excel values must not be silently lowercased
    if sheets:
        upper_issues: List[str] = []
        for sh in sheets:
            for row in sh['data_rows'][:30]:
                for cell in row:
                    if not isinstance(cell, str):
                        continue
                    val = cell.strip()
                    if val.isupper() and len(val) > 3 and val not in sgml:
                        # Check if a lowercased version appears instead
                        if val.lower() in sgml or val.capitalize() in sgml:
                            upper_issues.append(
                                f'"{val}" → appears lowercased in SGML')
                    if len(upper_issues) >= 5:
                        break
        if upper_issues:
            add('L4-D', 'HIGH',
                f'{len(upper_issues)} uppercase Excel value(s) '
                f'appear lowercased in SGML',
                detail=upper_issues)

    # L4-E: POLIDENT <N> and <DATE> must be non-empty (POLIDOC types only)
    if DOC_TYPES.get(doc_type, {}).get('root_tag') == 'POLIDOC':
        n_m    = re.search(r'<N>(.*?)</N>', sgml)
        date_m = re.search(r'<DATE>(.*?)</DATE>', sgml)
        if n_m and not n_m.group(1).strip():
            ln = _find_line(_lines, r'<N>')
            add('L4-E', 'HIGH', '<N> in POLIDENT is empty',
                line=ln,
                fix='Populate <N></N> with the document number from Excel metadata')
        if date_m and not date_m.group(1).strip():
            ln = _find_line(_lines, r'<DATE>')
            add('L4-E', 'HIGH', '<DATE> in POLIDENT is empty',
                line=ln,
                fix='Populate <DATE></DATE> with the effective date (Month D, YYYY)')

    # ── L4-F: <DATE> format must be "Month D, YYYY" ───────────────────────────
    date_m = re.search(r'<DATE>(.*?)</DATE>', sgml)
    if date_m and date_m.group(1).strip():
        date_val = date_m.group(1).strip()
        if not _DATE_RE.match(date_val):
            ln = _find_line(_lines, r'<DATE>')
            add('L4-F', 'MEDIUM',
                f'<DATE> format unexpected: "{date_val}" — '
                f'expected "Month D, YYYY" (e.g. "April 15, 2026")',
                line=ln,
                fix=f'Change <DATE>{date_val}</DATE> to '
                    f'<DATE>Month D, YYYY</DATE> format')

    return {'score': max(0, score), 'max': 10, 'issues': issues}


# ─── DECISION ENGINE ─────────────────────────────────────────────────────────
def make_decision(normalised: float, l2_score: int,
                  has_critical: bool) -> str:
    if has_critical or l2_score < 20:
        return 'REJECT'
    if normalised >= 90:
        return 'ACCEPT'
    if normalised >= 80:
        return 'ACCEPT_WITH_WARNINGS'
    if normalised >= 70:
        return 'REVIEW'
    return 'REJECT'


# ─── ACTIONABLE FIXES (for HITL app) ────────────────────────────────────────
def generate_excel_fixes(sgml: str, report: dict) -> list:
    """
    Return a sorted list of actionable fix dicts from a validate() report.

    Each fix dict contains:
      layer, check_id, severity, description, line (1-based, 0=unknown),
      context_before, suggested_fix, detail, auto_fixable, _fix_old, _fix_new
    """
    SEV_RANK      = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'INFO': 4}
    AUTO_CHECKS   = {'C11', 'C12', 'C16', 'C19', 'L4-A', 'L4-B'}
    fixes: list   = []
    _lines        = sgml.splitlines()

    for layer in ('L1', 'L2', 'L3', 'L4'):
        for iss in report.get('issues', {}).get(layer, []):
            sev      = iss.get('sev', 'LOW')
            check_id = iss.get('id', '')
            line_no  = iss.get('line', 0)

            auto    = check_id in AUTO_CHECKS
            old_str = new_str = ''

            if auto and line_no and 0 < line_no <= len(_lines):
                orig = _lines[line_no - 1]
                if check_id == 'C11':          # None → empty cell
                    new = orig.replace('>None<', '><')
                    if new != orig:
                        old_str, new_str = orig, new
                elif check_id == 'C16':        # accented chars → entities
                    new = orig
                    _ENT = {
                        'é': '&eacute;', 'è': '&egrave;', 'ê': '&ecirc;',
                        'à': '&agrave;', 'â': '&acirc;',  'ô': '&ocirc;',
                        'î': '&icirc;',  'ç': '&ccedil;', 'œ': '&oelig;',
                        'Œ': '&OElig;',
                    }
                    for char, ent in _ENT.items():
                        new = new.replace(char, ent)
                    if new != orig:
                        old_str, new_str = orig, new

            fixes.append({
                'layer':          layer,
                'check_id':       check_id,
                'severity':       sev,
                'description':    iss.get('msg', ''),
                'line':           line_no,
                'context_before': iss.get('context_before', ''),
                'suggested_fix':  iss.get('suggested_fix', ''),
                'detail':         iss.get('detail', []),
                'auto_fixable':   auto and bool(old_str),
                '_fix_old':       old_str,
                '_fix_new':       new_str,
            })

    fixes.sort(key=lambda f: (
        SEV_RANK.get(f['severity'], 9), f['layer'],
        0 if f['line'] > 0 else 1,
    ))
    return fixes


def apply_excel_fixes(sgml: str, fixes: list) -> tuple:
    """Apply all auto_fixable fixes in one pass. Returns (corrected_sgml, count)."""
    corrected = sgml
    applied   = 0
    for fix in fixes:
        if fix.get('auto_fixable') and fix.get('_fix_old') and fix.get('_fix_new'):
            if fix['_fix_old'] in corrected:
                corrected = corrected.replace(fix['_fix_old'], fix['_fix_new'], 1)
                applied  += 1
    return corrected, applied


# ─── MAIN VALIDATE ───────────────────────────────────────────────────────────
def validate(sgm_path: Path, xlsx_path: Optional[Path]) -> dict:
    sgml  = sgm_path.read_text(encoding='utf-8', errors='replace')
    lines = sgml.splitlines()
    xl_data = (load_excel(xlsx_path) if xlsx_path
               else {'error': 'No Excel file provided', 'sheets': []})

    doc_type = detect_doc_type(sgml)

    l1 = run_l1(sgml, xl_data, doc_type, lines)
    l2 = run_l2(sgml, doc_type, lines)
    l3 = run_l3(sgml, doc_type, lines)
    l4 = run_l4(sgml, xl_data, doc_type, lines)

    raw     = l1['score'] + l2['score'] + l3['score'] + l4['score']
    raw_max = l1['max']   + l2['max']   + l3['max']   + l4['max']   # 100
    normalised = round((raw / raw_max) * 100, 1)

    all_issues   = (l1['issues'] + l2['issues'] +
                    l3['issues'] + l4['issues'])
    has_critical = any(i.get('sev') == 'CRITICAL' for i in all_issues)
    decision     = make_decision(normalised, l2['score'], has_critical)

    return {
        'file':         sgm_path.name,
        'xlsx':         xlsx_path.name if xlsx_path else None,
        'doc_type':     doc_type,
        'doc_label':    DOC_TYPES.get(doc_type, {}).get('label', 'Unknown'),
        'timestamp':    datetime.now().isoformat(timespec='seconds'),
        'scores': {
            'L1_source_fidelity':    {'score': l1['score'], 'max': l1['max']},
            'L2_structural':         {'score': l2['score'], 'max': l2['max']},
            'L3_doctype_compliance': {'score': l3['score'], 'max': l3['max']},
            'L4_data_integrity':     {'score': l4['score'], 'max': l4['max']},
            'total':      raw,
            'total_max':  raw_max,
            'normalised': normalised,
        },
        'decision':     decision,
        'has_critical': has_critical,
        'issues': {
            'L1': l1['issues'],
            'L2': l2['issues'],
            'L3': l3['issues'],
            'L4': l4['issues'],
        },
        'stats': {
            'cell_count': l2.get('cell_count', 0),
            'xl_sheets':  xl_data.get('sheet_count', 0),
        },
    }


# ─── CONSOLE REPORT ──────────────────────────────────────────────────────────
_DECISION_BADGE = {
    'ACCEPT':               'ACCEPT',
    'ACCEPT_WITH_WARNINGS': 'ACCEPT WITH WARNINGS',
    'REVIEW':               'REVIEW (human check needed)',
    'REJECT':               'REJECT',
}
_SEV_LABEL = {
    'CRITICAL': '[CRIT]',
    'HIGH':     '[HIGH]',
    'MEDIUM':   '[MED] ',
    'LOW':      '[LOW] ',
    'INFO':     '[INFO]',
}


def print_report(r: dict) -> None:
    W   = 72
    sep = '-' * W
    thick = '=' * W
    sc  = r['scores']

    print()
    print(thick)
    print(f"  Excel -> SGML Validator               {r['timestamp']}")
    print(sep)
    print(f"  SGML  : {r['file']}")
    print(f"  Excel : {r['xlsx'] or '(not provided)'}")
    print(f"  Type  : {r['doc_label']}")
    print(sep)
    print(f"  L1 Source Fidelity     "
          f"{sc['L1_source_fidelity']['score']:>3}/"
          f"{sc['L1_source_fidelity']['max']}")
    print(f"  L2 Structural          "
          f"{sc['L2_structural']['score']:>3}/"
          f"{sc['L2_structural']['max']}")
    print(f"  L3 Doc-Type Compliance "
          f"{sc['L3_doctype_compliance']['score']:>3}/"
          f"{sc['L3_doctype_compliance']['max']}")
    print(f"  L4 Data Integrity      "
          f"{sc['L4_data_integrity']['score']:>3}/"
          f"{sc['L4_data_integrity']['max']}")
    print(sep)
    print(f"  TOTAL    {sc['total']}/{sc['total_max']}  "
          f"({sc['normalised']}%)")
    print(f"  DECISION --> {_DECISION_BADGE.get(r['decision'], r['decision'])}")
    print(sep)

    all_issues = []
    for layer in ('L1', 'L2', 'L3', 'L4'):
        for iss in r['issues'].get(layer, []):
            all_issues.append((layer, iss))

    crits  = [(l, i) for l, i in all_issues if i.get('sev') == 'CRITICAL']
    others = [(l, i) for l, i in all_issues if i.get('sev') != 'CRITICAL']

    if not all_issues:
        print('  No issues found — output looks clean.')
    else:
        if crits:
            print(f'  CRITICAL issues ({len(crits)}):')
            for layer, iss in crits:
                loc = f' (line {iss["line"]})' if iss.get('line') else ''
                print(f'    [{layer}] {_SEV_LABEL["CRITICAL"]} '
                      f'{iss["id"]}: {iss["msg"]}{loc}')
                for d in iss.get('detail', [])[:3]:
                    print(f'             -> {d}')
                if iss.get('suggested_fix'):
                    print(f'           FIX: {iss["suggested_fix"]}')
        if others:
            print(f'  Other issues ({len(others)}):')
            for layer, iss in others:
                lbl = _SEV_LABEL.get(iss.get('sev', ''), '[?]  ')
                loc = f' (line {iss["line"]})' if iss.get('line') else ''
                print(f'    [{layer}] {lbl} '
                      f'{iss["id"]}: {iss["msg"]}{loc}')
                for d in iss.get('detail', [])[:2]:
                    print(f'             -> {d}')
                if iss.get('suggested_fix'):
                    print(f'           FIX: {iss["suggested_fix"]}')

    print(thick)
    print()


# ─── CLI ENTRY POINT ─────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description='Excel -> SGML Validator (deterministic, no LLM required)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            'Examples:\n'
            '  python excel_validator.py output.sgm source.xlsx\n'
            '  python excel_validator.py output.sgm'
            '  # structural checks only\n'
        ),
    )
    parser.add_argument('sgm',  help='Path to the generated .sgm SGML file')
    parser.add_argument('xlsx', nargs='?', default=None,
                        help='Path to the source .xlsx Excel file (optional)')
    parser.add_argument('--json', metavar='FILE', default=None,
                        help='JSON report path '
                             '(default: <sgm_stem>.validator.json)')
    parser.add_argument('--fixes', action='store_true',
                        help='Apply auto-fixable issues and write corrected '
                             'SGML to <sgm_stem>.fixed.sgm')
    args = parser.parse_args()

    sgm_path  = Path(args.sgm)
    xlsx_path = Path(args.xlsx) if args.xlsx else None

    if not sgm_path.exists():
        print(f'ERROR: SGML file not found: {sgm_path}', file=sys.stderr)
        sys.exit(2)
    if xlsx_path and not xlsx_path.exists():
        print(f'[WARNING] Excel file not found: {xlsx_path} '
              f'— L1/L4 Excel checks will be skipped',
              file=sys.stderr)
        xlsx_path = None

    result = validate(sgm_path, xlsx_path)
    print_report(result)

    # Write JSON report
    json_out = (Path(args.json) if args.json
                else sgm_path.with_suffix('.validator.json'))
    json_out.write_text(
        json.dumps(result, indent=2, ensure_ascii=False), encoding='utf-8')
    print(f'  JSON report -> {json_out}')

    # --fixes: apply auto-fixable issues
    if args.fixes:
        sgml_text = sgm_path.read_text(encoding='utf-8', errors='replace')
        fixes     = generate_excel_fixes(sgml_text, result)
        corrected, n_applied = apply_excel_fixes(sgml_text, fixes)
        if n_applied:
            fixed_out = sgm_path.with_suffix('.fixed.sgm')
            fixed_out.write_text(corrected, encoding='utf-8')
            print(f'  Auto-fixed {n_applied} issue(s) -> {fixed_out}')
        else:
            print('  No auto-fixable issues found.')
    print()

    # Exit codes: 0=ACCEPT/ACCEPT_WITH_WARNINGS, 1=REVIEW, 2=REJECT
    sys.exit({'ACCEPT': 0, 'ACCEPT_WITH_WARNINGS': 0,
              'REVIEW': 1, 'REJECT': 2}.get(result['decision'], 2))


if __name__ == '__main__':
    main()

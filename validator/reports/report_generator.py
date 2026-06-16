"""
reports/report_generator.py
────────────────────────────
Generate structured validation reports in JSON and Excel formats.

Used by the test harness and the main validator pipeline.
"""

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, Union, TYPE_CHECKING

if TYPE_CHECKING:
    from validator.validator_main import ValidationReport


def save_json_report(report: "ValidationReport", output_path: str) -> None:
    """Write a ValidationReport to a JSON file."""
    data = report.to_dict()
    data["generated_at"] = datetime.utcnow().isoformat() + "Z"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def save_excel_report(
    reports: list["ValidationReport"],
    output_path: str,
) -> None:
    """
    Write a list of ValidationReports to an Excel workbook.

    Requires openpyxl.  If not available, falls back to CSV.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        _write_excel(reports, output_path)
    except ImportError:
        csv_path = str(output_path).replace(".xlsx", ".csv")
        _write_csv(reports, csv_path)
        print(f"openpyxl not installed — wrote CSV to {csv_path}")


def _write_excel(reports: list, output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Results"

    # Header
    headers = [
        "SGML File", "PDF File", "Decision",
        "Total (100)", "L1 (35)", "L2 (40)", "L3 (25)",
        "L1 Text", "L1 Section", "L1 Table", "L1 Footnote",
        "L2 Schema", "L2 Nesting", "L2 Entity", "L2 Table", "L2 Graphics",
        "L2 Content", "L2 Legal",
        "L3 Jurisdiction", "L3 Statistical", "L3 Pattern",
        "Jurisdiction", "Doc Type",
        "Critical Failures", "Issues Count", "Error",
    ]
    ws.append(headers)

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Decision colors
    DECISION_COLORS = {
        "ACCEPT": "C6EFCE",
        "ACCEPT_WITH_WARNINGS": "FFEB9C",
        "REVIEW": "FFEB9C",
        "REJECT": "FFC7CE",
    }

    for report in reports:
        d = report.to_dict()
        scores = d.get("scores", {})
        l1d = d.get("l1_details") or {}
        l2d = d.get("l2_details") or {}
        l3d = d.get("l3_details") or {}

        row = [
            Path(d["sgml_path"]).name if d["sgml_path"] else "",
            Path(d["pdf_path"]).name if d["pdf_path"] else "",
            d.get("decision", ""),
            scores.get("total", ""),
            scores.get("l1_content_fidelity", ""),
            scores.get("l2_structural", ""),
            scores.get("l3_corpus_pattern", ""),
            l1d.get("text_completeness_score", ""),
            l1d.get("section_completeness_score", ""),
            l1d.get("table_completeness_score", ""),
            l1d.get("footnote_completeness_score", ""),
            l2d.get("schema_score", ""),
            l2d.get("nesting_score", ""),
            l2d.get("entity_score", ""),
            l2d.get("table_score", ""),
            l2d.get("graphics_score", ""),
            l2d.get("content_rules_score", ""),
            l2d.get("legal_structure_score", ""),
            l3d.get("jurisdiction_score", ""),
            l3d.get("statistical_score", ""),
            l3d.get("pattern_score", ""),
            l3d.get("detected_jurisdiction", ""),
            l3d.get("detected_doc_type", ""),
            "; ".join(d.get("critical_failures", [])),
            len(d.get("issues", [])),
            d.get("error") or "",
        ]
        ws.append(row)

        # Color the decision cell (column C = index 3)
        decision = d.get("decision", "")
        color = DECISION_COLORS.get(decision, "FFFFFF")
        fill = PatternFill("solid", fgColor=color)
        ws.cell(row=ws.max_row, column=3).fill = fill

    # Auto-fit columns
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 22

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Excel report written to {output_path}")


def _write_csv(reports: list, output_path: str) -> None:
    import csv

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "SGML File", "Decision", "Total", "L1", "L2", "L3",
            "Jurisdiction", "Doc Type", "Critical Failures", "Issues",
        ])
        for report in reports:
            d = report.to_dict()
            scores = d.get("scores", {})
            l3d = d.get("l3_details") or {}
            writer.writerow([
                Path(d["sgml_path"]).name if d["sgml_path"] else "",
                d.get("decision", ""),
                scores.get("total", ""),
                scores.get("l1_content_fidelity", ""),
                scores.get("l2_structural", ""),
                scores.get("l3_corpus_pattern", ""),
                l3d.get("detected_jurisdiction", ""),
                l3d.get("detected_doc_type", ""),
                "; ".join(d.get("critical_failures", [])),
                len(d.get("issues", [])),
            ])

    print(f"CSV report written to {output_path}")


def format_summary(reports: list, title: str = "Validation Summary") -> str:
    """Return a human-readable summary string."""
    from collections import Counter

    decisions = Counter(r.decision for r in reports)
    total = len(reports)
    passed = decisions.get("ACCEPT", 0) + decisions.get("ACCEPT_WITH_WARNINGS", 0)

    if total == 0:
        return f"{title}\n  No files validated."

    scores = [r.total_score for r in reports if r.total_score > 0]
    avg_score = sum(scores) / len(scores) if scores else 0

    lines = [
        f"\n{'='*60}",
        f"  {title}",
        f"{'='*60}",
        f"  Total files:   {total}",
        f"  Passed:        {passed}  ({100*passed/total:.1f}%)",
        f"  Review:        {decisions.get('REVIEW', 0)}",
        f"  Rejected:      {decisions.get('REJECT', 0)}",
        f"  Average score: {avg_score:.1f}/100",
        f"{'='*60}",
        f"  Decisions:",
        *[f"    {k}: {v}" for k, v in decisions.most_common()],
        f"{'='*60}\n",
    ]
    return "\n".join(lines)
